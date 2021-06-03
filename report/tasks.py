from openpyxl.reader.excel import load_workbook
from .scripts.stockReport import WebSheets, startpoint, SAPDump, WebSheets, RRP1Copy, SAPRRP1ToWorking
from celery.task.schedules import crontab
from celery.decorators import periodic_task, task
from celery import shared_task
from .scripts.sapPws import pwsStart
from celery_progress.backend import ProgressRecorder
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import datetime

@shared_task
def run_pws():
    # print('nice')
    result = pwsStart()

@periodic_task(run_every=(crontab(hour=8, minute=30)), ignore_result=True)
def pws_task():
    run_pws.delay()

@shared_task(bind=True, name="logic", ignore_result=False)
def logic(self, variant):
    # filename = "x"
    # sleep(10)
    # progress_recorder = ProgressRecorder(self)
    # for i in range(0,1):
    #     res = startpoint(variant)
    #     progress_recorder.set_progress(i, 1, description="Processing")
    progress_recorder = ProgressRecorder(self)
    print('Started')
    if variant=="paste":
        wb1 = load_workbook('report/input/Base File (Paste).xlsx')
    elif variant=="brush":
        wb1 = load_workbook('report/input/Base File (Brush).xlsx')
    else:
        wb1 = load_workbook('report/input/Base File (PCP).xlsx')
    ws1 = wb1.active

    # ws2 = wb1.worksheets[1]

    # ws3 = wb1.worksheets[2]
    SAPDump(wb1, variant)

    progress_recorder.set_progress(1, 5, description="Processed PWS Report")

    # WebSheets(wb1, "", "report/input/web_south.xlsx", 3)
    # WebSheets(wb1, "", "report/input/web_north.xlsx", 4)
    # WebSheets(wb1, "", "report/input/web_west.xlsx", 5)
    # WebSheets(wb1, "", "report/input/web_east.xlsx", 6)

    # web download
    WebSheets(wb1, "http://cpindia.win.colpal.com/hubli/scripts/dwnld_xls.asp", "report/input/web_south", 3)
    WebSheets(wb1, "http://cpindia.win.colpal.com/north/scripts/dwnld_xls.asp", "report/input/web_north", 4)
    WebSheets(wb1, "http://cpindia.win.colpal.com/despatch/scripts/dwnld_xls.asp", "report/input/web_west", 5)
    WebSheets(wb1, "http://cpindia.win.colpal.com/koldc/scripts/dwnld_xls.asp", "report/input/web_east", 6)
    
    progress_recorder.set_progress(2, 5, description="Downloaded DC Input")
    RRP1Copy(wb1)
    
    progress_recorder.set_progress(3, 5, description="Processed RRP1 Input")
    SAPRRP1ToWorking(wb1)
    
    progress_recorder.set_progress(4, 5, description="Report Generated")
    # ParentChildSku(wb1)
    for row in ws1.iter_rows(min_row= 1, max_row = ws1.max_row, min_col = 1, max_col = ws1.max_column):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    ws7 = wb1.worksheets[7]
    ws7.delete_cols(15)
    ws7.delete_cols(15)
    filename = datetime.date.today().strftime("%d%b%Y")
    if variant=="paste":
        wb1.save('report/output/'+filename+' (Paste).xlsx')
    elif variant=="brush":
        wb1.save('report/output/'+filename+' (Brush).xlsx')
    else:
        wb1.save('report/output/'+filename+' (PCP).xlsx')
    
    progress_recorder.set_progress(5, 5, description="Report Saved")
    # return True