from openpyxl.workbook import workbook
from pyrfc import Connection, ABAPApplicationError, ABAPRuntimeError, LogonError, CommunicationError
from openpyxl import Workbook

def getPwsData():
    try:
        conn = Connection(user='INBHP002', ashost='CAIapp07.esc.win.colpal.com', sysnr='07', client='321', passwd='Bdp@251299')
        print(conn.alive)
        plant = [{'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN08'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN15'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN16'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN59'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN60'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN61'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN62'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN63'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN64'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN65'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN66'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN67'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN68'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN69'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN70'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN71'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN72'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN73'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN74'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN75'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN76'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN77'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN78'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN79'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN80'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN81'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN82'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN83'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN84'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN85'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN86'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN87'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN88'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN89'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN90'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN91'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN95'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN97'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN98'},
        {'SIGN':'I',  'OPTION':'EQ', 'LOW':'IN99'}]

        material = [{'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600906'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600908'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600917'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600931'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600933'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600951'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1600958'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601002'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601013'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601025'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601032'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601035'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601052'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601108'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601226'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601230'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601240'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601261'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601288'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601291'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1601989'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602104'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602154'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602265'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602266'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602269'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602270'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602601'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602881'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1602891'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607704'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607705'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607707'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607714'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607715'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607724'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607750'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607778'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607779'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607790'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607798'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607799'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607805'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1607821'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1608101'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1608102'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1608103'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1608104'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1608125'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1611441'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1611442'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1611443'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1611451'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'1611452'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'BIN00676A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00549A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00554A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00557A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00558A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00559A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00560A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00597A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00598A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00599A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00604A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00605A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00606A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00616A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00618A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00619A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00641A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00642A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00650A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00651A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00653A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00654A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00655A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00659A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00670A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00676A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00697A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00704A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00705A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00712A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00747A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00749A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00751A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00756A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00757A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00761A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00763A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00771A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00800A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00806A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00808A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00809A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00810A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00811A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00822A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00825A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00829A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00831A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00832A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00833A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00834A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00835A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00836A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00837A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00838A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00854A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00856A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00857A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00858A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00867A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00868A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00875A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00894A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'TH01876A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'TH02054A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00934A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00932A'},
        {'SIGN':'I', 'OPTION':'EQ', 'LOW':'IN00872A'}]
        
        fm_dict = {
            'function_name':'Z1A_PWS_INDIA',
            'connection': conn,
            'import_args': {
                'IT_PLANT': plant,
                'IT_MATERIAL': material
            }
        }
        # result = conn.call("Z1A_PWS_INDIA", IT_PLANT = plant, IT_MATERIAL = material)
        result = fm_dict['connection'].call(fm_dict['function_name'], **fm_dict['import_args'])
        # print(result['ET_EXCELTAB'])
        resultTable = result['ET_EXCELTAB']

        conn.close()
    except CommunicationError:
        print(u"Could not connect to server.")
        raise
    except LogonError:
        print(u"Could not log in. Wrong credentials?")
        raise
    except (ABAPApplicationError, ABAPRuntimeError):
        print(u"An error occurred.")
        raise
    finally:
        conn.close()

    # test dict
    testdict = {
        'MATNR': '0000000000AS16A1013', 'MAKTX': 'CDC TP 200+100+Tb', 'WERKS': 'IN95', 'GSBEST': '        0', 'MZUBB': '        0', 'IN_TRANSIT': '     1221', 'TOTPOTENTIAL': '        0', 'FKIMG': '        0', 
        'LOSTSALES': ' 0', 'ON_HAND': '     4816', 'NETAVAIL': '     6037', 'ALLOC_QTY': '        0', 'BAL_TO_DESP': '        0', 'EXCESS_COV': '     6037', 'EISBE': '     1778', 'MINBE': '     5337', 'BSTMI': '     2810', 'BSTMIL': '        0'
    }
    # resultTable.append(testdict)
    return resultTable

# save to excel
def saveExcel(resultTable):
    wb1 = Workbook()
    ws1 = wb1.active
    x = 1
    for i in range(len(resultTable)):
        if i!=0:
            if resultTable[i]['MATNR'].isdigit():
                ws1.cell(row = i+1, column = 1).value = int(resultTable[i]['MATNR'])
            else:
                ws1.cell(row = i+1, column = 1).value = resultTable[i]['MATNR']
            ws1.cell(row = i+1, column = 2).value = resultTable[i]['MAKTX']
            ws1.cell(row = i+1, column = 3).value = resultTable[i]['WERKS']
            ws1.cell(row = i+1, column = 4).value = int(resultTable[i]['GSBEST'])
            ws1.cell(row = i+1, column = 5).value = int(resultTable[i]['MZUBB'])
            ws1.cell(row = i+1, column = 6).value = int(resultTable[i]['IN_TRANSIT'])
            ws1.cell(row = i+1, column = 7).value = int(resultTable[i]['TOTPOTENTIAL'])
            ws1.cell(row = i+1, column = 8).value = int(resultTable[i]['FKIMG'])
            ws1.cell(row = i+1, column = 9).value = int(resultTable[i]['LOSTSALES'])
            ws1.cell(row = i+1, column = 10).value = int(resultTable[i]['ON_HAND'])
            ws1.cell(row = i+1, column = 11).value = int(resultTable[i]['NETAVAIL'])
            ws1.cell(row = i+1, column = 12).value = int(resultTable[i]['ALLOC_QTY'])
            ws1.cell(row = i+1, column = 13).value = int(resultTable[i]['BAL_TO_DESP'])
            ws1.cell(row = i+1, column = 14).value = int(resultTable[i]['EXCESS_COV'])
            ws1.cell(row = i+1, column = 15).value = int(resultTable[i]['EISBE'])
            ws1.cell(row = i+1, column = 16).value = int(resultTable[i]['MINBE'])
            ws1.cell(row = i+1, column = 17).value = int(resultTable[i]['BSTMI'])
            ws1.cell(row = i+1, column = 18).value = int(resultTable[i]['BSTMIL'])
        else:
            ws1.cell(row = i+1, column = 1).value = resultTable[i]['MATNR']
            ws1.cell(row = i+1, column = 2).value = resultTable[i]['MAKTX']
            ws1.cell(row = i+1, column = 3).value = resultTable[i]['WERKS']
            ws1.cell(row = i+1, column = 4).value = resultTable[i]['GSBEST']
            ws1.cell(row = i+1, column = 5).value = resultTable[i]['MZUBB']
            ws1.cell(row = i+1, column = 6).value = resultTable[i]['IN_TRANSIT']
            ws1.cell(row = i+1, column = 7).value = resultTable[i]['TOTPOTENTIAL']
            ws1.cell(row = i+1, column = 8).value = resultTable[i]['FKIMG']
            ws1.cell(row = i+1, column = 9).value = resultTable[i]['LOSTSALES']
            ws1.cell(row = i+1, column = 10).value = resultTable[i]['ON_HAND']
            ws1.cell(row = i+1, column = 11).value = resultTable[i]['NETAVAIL']
            ws1.cell(row = i+1, column = 12).value = resultTable[i]['ALLOC_QTY']
            ws1.cell(row = i+1, column = 13).value = resultTable[i]['BAL_TO_DESP']
            ws1.cell(row = i+1, column = 14).value = resultTable[i]['EXCESS_COV']
            ws1.cell(row = i+1, column = 15).value = resultTable[i]['EISBE']
            ws1.cell(row = i+1, column = 16).value = resultTable[i]['MINBE']
            ws1.cell(row = i+1, column = 17).value = resultTable[i]['BSTMI']
            ws1.cell(row = i+1, column = 18).value = resultTable[i]['BSTMIL']

    wb1.save("report/input/pws.xlsx")
    return True

def pwsStart():
    resultTable = getPwsData()
    res = saveExcel(resultTable)
    return res