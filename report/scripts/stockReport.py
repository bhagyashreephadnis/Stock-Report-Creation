import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from xls2xlsx import XLS2XLSX
import datetime, os, requests

def SAPDump(wb1, variant):
    # x2x = XLS2XLSX("AUTOCAP20210217020006.XLS")
    # x2x.to_xlsx("AUTOCAP20210217020006.xlsx")
    if variant=="paste":
        wb2 = load_workbook('report/input/pws_paste.xlsx')
    elif variant=="brush":
        wb2 = load_workbook('report/input/pws_brush.xlsx')
    else:
        wb2 = load_workbook('report/input/pws_other.xlsx')
    ws1 = wb2.worksheets[0]
    mr = ws1.max_row
    mc = ws1.max_column

    ws2 = wb1.worksheets[2]
    k = 1
    for i in range (1, mr + 1):
        c = ws1.cell(row = i, column = 1)
        if c.value is None:
            continue
        for j in range (1, mc + 1):
            c = ws1.cell(row = i, column = j)
            if isinstance(c.value, str) and c.value[-1]=='-':
                ws2.cell(row = k, column = j).value = int(c.value[:-1])
            else:
                ws2.cell(row = k, column = j).value = c.value
        k += 1

def WebSheets(wb1, url, filename, sheet_number):
    page = requests.get(url)
    open(filename+".xls", "wb").write(page.content)

    x2x = XLS2XLSX(filename+".xls")
    x2x.to_xlsx(filename+".xlsx")
    os.remove(filename+".xls")

    wb2 = load_workbook(filename+".xlsx")
    # wb2 = load_workbook(filename)
    ws1 = wb2.worksheets[0]
    mr = ws1.max_row
    mc = ws1.max_column

    ws2 = wb1.worksheets[sheet_number]
    if mr==2:
        for i in range(2):
            for j in range(1, mc+1):
                if j==5 and i==0:
                    ws2.cell(row = i+1, column = j).value = "Concatenate"
                elif j>5 and i==0:
                    ws2.cell(row = i+1, column = j).value = ws1.cell(row = i+1, column = j-1).value
                else:
                    ws2.cell(row = i+1, column = j).value = ws1.cell(row = i+1, column = j).value
        # ws2.insert_cols(5)
        # ws2.cell(row = 1, column = 5).value = "Concatenate"
    else:
        for i in range (1, mr + 1):
            c = ws1.cell(row = i, column = 1)
            if c.value is None:
                continue
            for j in range (1, mc + 2):
                if j==5:
                    if i==1:
                        ws2.cell(row = i, column = 5).value = "Concatenate"
                    else:
                        c = str(ws1.cell(row = i, column = 2).value) + ws1.cell(row =i , column = 4).value
                        ws2.cell(row = i, column = 5).value = c
                elif j>5:
                    c = ws1.cell(row = i, column = j-1)
                    ws2.cell(row = i, column = j).value = c.value
                else:
                    c = ws1.cell(row = i, column = j)
                    ws2.cell(row = i, column = j).value = c.value

def RRP1Copy(wb1):
    wb2 = load_workbook('report/input/rrp1.xlsx')
    ws1 = wb2.worksheets[0]
    mr = ws1.max_row
    mc = ws1.max_column

    ws2 = wb1.worksheets[7]
    k = 2

    ws2.insert_cols(15)
    ws2.cell(row=1, column=15).value = "Sum of Quantity"
    ws2.cell(row=2, column=15).value = "Combi"
    ws2.insert_cols(16)
    ws2.cell(row=2, column=16).value = "Total"
    values = []
    l = 3
    for i in range (2, mr + 1):
        c = ws1.cell(row = i, column = 3)
        if c.value=="ConRel" or c.value=="StkTrsfDel":
            for j in range (1, mc + 3):
                if j==6:
                    ws2.cell(row = k, column = j).value = str(ws1.cell(row = i, column = 1).value) + str(ws1.cell(row = i, column = 5).value) + str(ws1.cell(row = i, column = 6).value)
                elif j==7:
                    ws2.cell(row = k, column = j).value = int(ws1.cell(row = i, column = 7).value) * -1
                elif j>7 and j<=9:
                    c = ws1.cell(row = i, column = j-2)
                    ws2.cell(row = k, column = j).value = c.value
                elif j==10:
                    ws2.cell(row = k, column = j).value = ws1.cell(row = i, column = j-2).value.date()
                elif j==11:
                    ws2.cell(row = k, column = j).value = ws1.cell(row = i, column = j-2).value.strftime("%I:%M:%S %p")
                elif j==12:
                    if ws1.cell(row = i, column = j-2).value is not None:
                        ws2.cell(row = k, column = j).value = ws1.cell(row = i, column = j-2).value.date()
                else:
                    c = ws1.cell(row = i, column = j)
                    ws2.cell(row = k, column = j).value = c.value
            # k += 1
            if ws2.cell(row = k, column = 6).value not in values:
                ws2.cell(row = l, column = 15).value = ws2.cell(row = k, column = 6).value
                ws2.cell(row = l, column = 16).value = ws2.cell(row = k, column = 7).value
                values.append(ws2.cell(row = k, column = 6).value)
                l += 1
            else:
                for x in range(3, i+1):
                    if ws2.cell(row = k, column = 6).value==ws2.cell(row = x, column = 15).value:
                        ws2.cell(row = x, column = 16).value += ws2.cell(row = k, column = 7).value
                        break
            k += 1

def SAPRRP1ToWorking(wb1):
    ws1 = wb1.worksheets[0] #working
    ws2 = wb1.worksheets[14] #warehouse codes
    ws3 = wb1.worksheets[2] #sap dump
    ws4 = wb1.worksheets[11] #next months forecast
    ws5 = wb1.worksheets[10] #old working
    ws8 = wb1.worksheets[8] #sku_mast
    # mr = ws1.max_row
    # mr2 = ws2.max_row
    mr3 = ws3.max_row
    
    c = 2
    parentval = list(cell.value for cell in ws8['A'] if cell.value is not None)[1:]
    childval = list(cell.value for cell in ws8['B'] if cell.value is not None)[1:]
    # print(parentval, childval)
    parentdict = dict((l, [0, 0]) for l in parentval)
    childdict = dict((l, [0, 0]) for l in childval)
    # print(parentdict, childdict)

    for i in range(2, mr3+1):
    # for i in range(2, 4):
        if ws3.cell(row = i, column = 1).value=="Mtot":
            ws1.cell(row = i, column = 1).value = "------------"
            # styling
            ws1.cell(row = i, column = 4).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
            ws1.cell(row = i, column = 6).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
            ws1.cell(row = i, column = 8).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D8E4BC'))
            ws1.cell(row = i, column = 10).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))
            ws1.cell(row = i, column = 12).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C4D79B'))
            ws1.cell(row = i, column = 13).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))
            ws1.cell(row = i, column = 15).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='0095B3D7'))
            ws1.cell(row = i, column = 16).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
            ws1.cell(row = i, column = 18).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00A6A6A6'))
            ws1.cell(row = i, column = 19).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D9D9D9'))
            ws1.cell(row = i, column = 21).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00DCE6F1'))
            ws1.cell(row = i, column = 22).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
            ws1.cell(row = i, column = 24).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))
            ws1.cell(row = i, column = 25).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C4D79B'))
            ws1.cell(row = i, column = 27).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D9D9D9'))
            ws1.cell(row = i, column = 32).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))
            ws1.cell(row = i, column = 34).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))
            # parent child wh
            if i!=2:
                ParentChildWh(wb1, i-1, c-1)
            continue
        # working
        ws1.cell(row = i, column = 1).value = ws3.cell(row = i, column = 1).value
        ws1.cell(row = i, column = 2).value = ws3.cell(row = i, column = 2).value
        ws1.cell(row = i, column = 3).value = ws3.cell(row = i, column = 3).value
        ws1.cell(row = i, column = 4).value = ws3.cell(row = i, column = 6).value
        ws1.cell(row = i, column = 5).value = ws3.cell(row = i, column = 7).value
        ws1.cell(row = i, column = 6).value = ws3.cell(row = i, column = 8).value
        ws1.cell(row = i, column = 7).value = ws3.cell(row = i, column = 9).value
        ws1.cell(row = i, column = 8).value = ws3.cell(row = i, column = 10).value
        ws1.cell(row = i, column = 9).value = ws3.cell(row = i, column = 11).value
        ws1.cell(row = i, column = 10).value = ws3.cell(row = i, column = 12).value
        ws1.cell(row = i, column = 11).value =  ws1.cell(row = i, column = 4).value +  ws1.cell(row = i, column = 6).value +  ws1.cell(row = i, column = 8).value -  ws1.cell(row = i, column = 10).value
        ws1.cell(row = i, column = 28).value = ws3.cell(row = i, column = 15).value
        ws1.cell(row = i, column = 29).value = ws3.cell(row = i, column = 16).value
        ws1.cell(row = i, column = 30).value = ws3.cell(row = i, column = 17).value
        ws1.cell(row = i, column = 31).value = ws3.cell(row = i, column = 18).value

        # index for parent child sku
        if ws1.cell(row = i, column = 1).value in parentval and parentdict[ws1.cell(row = i, column = 1).value]==[0, 0]:
            parentdict[ws1.cell(row = i, column = 1).value][0] = i
            parentdict[ws1.cell(row = i, column = 1).value][1] = c
        if ws1.cell(row = i, column = 1).value in childval and childdict[ws1.cell(row = i, column = 1).value]==[0, 0]:
            childdict[ws1.cell(row = i, column = 1).value][0] = i
            childdict[ws1.cell(row = i, column = 1).value][1] = c
        
        # old working
        ws5.cell(row = c, column = 1).value = ws3.cell(row = i, column = 1).value
        ws5.cell(row = c, column = 2).value = ws3.cell(row = i, column = 2).value
        ws5.cell(row = c, column = 3).value = ws3.cell(row = i, column = 3).value
        ws5.cell(row = c, column = 4).value = ws3.cell(row = i, column = 4).value
        ws5.cell(row = c, column = 5).value = ws3.cell(row = i, column = 5).value
        ws5.cell(row = c, column = 6).value = ws3.cell(row = i, column = 6).value
        ws5.cell(row = c, column = 7).value = ws3.cell(row = i, column = 7).value
        ws5.cell(row = c, column = 8).value = ws3.cell(row = i, column = 8).value
        ws5.cell(row = c, column = 9).value = ws3.cell(row = i, column = 9).value
        ws5.cell(row = c, column = 10).value = ws3.cell(row = i, column = 10).value
        ws5.cell(row = c, column = 11).value = ws3.cell(row = i, column = 11).value
        ws5.cell(row = c, column = 12).value = ws3.cell(row = i, column = 12).value
        ws5.cell(row = c, column = 13).value = ws3.cell(row = i, column = 13).value
        ws5.cell(row = c, column = 14).value = ws5.cell(row = c, column = 6).value + ws5.cell(row = c, column = 8).value + ws5.cell(row = c, column = 10).value - ws5.cell(row = c, column = 12).value

        search1 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN99"
        ws1.cell(row = i, column = 12).value = RRP1Lookup(wb1, search1)

        search2 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN15"
        ws1.cell(row = i, column = 13).value = RRP1Lookup(wb1, search2)

        search3 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN16"
        ws1.cell(row = i, column = 14).value = RRP1Lookup(wb1, search3)

        search4 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN08"
        ws1.cell(row = i, column = 15).value = RRP1Lookup(wb1, search4)

        search5 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN95"
        val1 = RRP1Lookup(wb1, search5)
        val2 = WebLookup(wb1, search5[:-4], 4)
        ws1.cell(row = i, column = 16).value = RRP1PlusWeb(val1, val2)

        search6 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN97"
        val3 = RRP1Lookup(wb1, search6)
        val4 = WebLookup(wb1, search6[:-4], 3)
        ws1.cell(row = i, column = 17).value = RRP1PlusWeb(val3, val4)

        search7 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN82"
        val5= RRP1Lookup(wb1, search7)
        val6 = WebLookup(wb1, search7[:-4], 5)
        ws1.cell(row = i, column = 18).value = RRP1PlusWeb(val5, val6)

        search8 = str(ws1.cell(row = i, column = 1).value) + ws1.cell(row = i, column = 3).value + "IN98"
        val7 = RRP1Lookup(wb1, search8)
        val8 = WebLookup(wb1, search8[:-4], 6)
        ws1.cell(row = i, column = 19).value = RRP1PlusWeb(val7, val8)

        # OH Stock %
        if ws1.cell(row = i, column = 10).value!=0:
            ws1.cell(row = i, column = 24).value = str(int(ws1.cell(row = i, column = 8).value/ws1.cell(row = i, column = 10).value*100))+"%"
        
        # Invoice Stock %
        if ws1.cell(row = i, column = 10).value!=0:
            # print(i, ws1.cell(row = i, column = 6).value)
            ws1.cell(row = i, column = 26).value = str(int(ws1.cell(row = i, column = 6).value/ws1.cell(row = i, column = 10).value*100))+"%"
        else:
            ws1.cell(row = i, column = 26).value = "0%"
        
        # UR Stock %
        if ws1.cell(row = i, column = 8).value>0 and ws1.cell(row = i, column = 10).value>0:
            ws1.cell(row = i, column = 27).value = str(int((ws1.cell(row = i, column = 8).value-ws1.cell(row = i, column = 7).value)/ws1.cell(row = i, column = 10).value*100))+"%"
        
        # wh name & branch
        for z in range(2, ws2.max_row+1):
            if ws2.cell(row = z, column = 1).value!=None:
                if ws1.cell(row = i, column = 3).value==ws2.cell(row = z, column = 1).value:
                    # print('here')
                    ws1.cell(row = i, column = 22).value = ws2.cell(row = z, column = 2).value
                    ws1.cell(row = i, column = 23).value = ws2.cell(row = z, column = 3).value
                    ws5.cell(row = c, column = 18).value = ws2.cell(row = z, column = 2).value
                    ws5.cell(row = c, column = 19).value = ws2.cell(row = z, column = 3).value
                    break
        
        # N E W S
        if ws1.cell(row = i, column = 23).value=="North" or ws1.cell(row = i, column = 23).value=="North DC":
            ws1.cell(row = i, column = 32).value = sum([int(ws1.cell(row = i, column = t).value) if ws1.cell(row = i, column = t).value!="-" else 0 for t in range(12,20)])
        else:
            ws1.cell(row = i, column = 32).value = 0
        if ws1.cell(row = i, column = 23).value=="East" or ws1.cell(row = i, column = 23).value=="East DC":
            ws1.cell(row = i, column = 33).value = sum([int(ws1.cell(row = i, column = t).value) if ws1.cell(row = i, column = t).value!="-" else 0 for t in range(12,20)])
        else:
            ws1.cell(row = i, column = 33).value = 0
        if ws1.cell(row = i, column = 23).value=="West" or ws1.cell(row = i, column = 23).value=="West DC":
            ws1.cell(row = i, column = 34).value = sum([int(ws1.cell(row = i, column = t).value) if ws1.cell(row = i, column = t).value!="-" else 0 for t in range(12,20)])
        else:
            ws1.cell(row = i, column = 34).value = 0
        if ws1.cell(row = i, column = 23).value=="South" or ws1.cell(row = i, column = 23).value=="South DC":
            ws1.cell(row = i, column = 35).value = sum([int(ws1.cell(row = i, column = t).value) if ws1.cell(row = i, column = t).value!="-" else 0 for t in range(12,20)])
        else:
            ws1.cell(row = i, column = 35).value = 0
        
        # M+1 FC
        for z in range(2, ws4.max_row+1):
            if ws4.cell(row=z, column=5).value!=None:
                if str(str(ws1.cell(row = i, column = 1).value)+ws1.cell(row = i, column = 3).value)==str(ws4.cell(row = z, column = 1).value)+str(ws4.cell(row = z, column = 3).value):
                    ws1.cell(row = i, column = 25).value = round(ws4.cell(row = z, column = 5).value)
                    break
        if ws1.cell(row = i, column = 25).value is None:
            ws1.cell(row = i, column = 25).value = 0
        ws5.cell(row = c, column = 26).value = ws1.cell(row = i, column = 25).value
        
        # coverage cast
        sumval = sum([int(ws1.cell(row = i, column = t).value) if ws1.cell(row = i, column = t).value!="-" else 0 for t in range(12,20)])
        # print(sumval)
        sumval += ws1.cell(row = i, column = 4).value+ws1.cell(row = i, column = 6).value+ws1.cell(row = i, column = 8).value
        if ws1.cell(row = i, column = 10).value!=0 and sumval <= ws1.cell(row = i, column = 10).value:
            ws1.cell(row = i, column = 20).value = str(round(sumval/ws1.cell(row = i, column = 10).value*100))+"%"
            ws5.cell(row = c, column = 16).value = ws1.cell(row = i, column = 20).value
        elif ws1.cell(row = i, column = 25).value!=0:
            ws1.cell(row = i, column = 20).value = str(round(100+((sumval - ws1.cell(row = i, column = 10).value)/ws1.cell(row = i, column = 25).value*100)))+"%"
            ws5.cell(row = c, column = 16).value = ws1.cell(row = i, column = 20).value

        # modified
        if ws1.cell(row = i, column = 10).value!=0 and ws1.cell(row = i, column = 4).value+ws1.cell(row = i, column = 6).value+ws1.cell(row = i, column = 8).value<=ws1.cell(row = i, column = 10).value:
            ws1.cell(row = i, column = 21).value = str(round((ws1.cell(row = i, column = 4).value+ws1.cell(row = i, column = 6).value+ws1.cell(row = i, column = 8).value)/ws1.cell(row = i, column = 10).value*100))+"%"
        elif ws1.cell(row = i, column = 25).value!=0:
            ws1.cell(row = i, column = 21).value = str(round(100 + (ws1.cell(row = i, column = 4).value+ws1.cell(row = i, column = 6).value+ws1.cell(row = i, column = 8).value-ws1.cell(row = i, column = 10).value)/ws1.cell(row = i, column = 25).value*100))+"%"
        
        # styling
        ws1.cell(row = i, column = 4).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
        ws1.cell(row = i, column = 6).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
        ws1.cell(row = i, column = 8).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D8E4BC'))
        ws1.cell(row = i, column = 10).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))
        ws1.cell(row = i, column = 12).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C4D79B'))
        ws1.cell(row = i, column = 13).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))
        ws1.cell(row = i, column = 15).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='0095B3D7'))
        ws1.cell(row = i, column = 16).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
        ws1.cell(row = i, column = 18).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00A6A6A6'))
        ws1.cell(row = i, column = 19).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D9D9D9'))
        ws1.cell(row = i, column = 21).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00DCE6F1'))
        ws1.cell(row = i, column = 22).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
        ws1.cell(row = i, column = 24).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))
        ws1.cell(row = i, column = 25).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C4D79B'))
        ws1.cell(row = i, column = 27).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D9D9D9'))
        ws1.cell(row = i, column = 32).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))
        ws1.cell(row = i, column = 34).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00EBF1DE'))

        c += 1
    
    # print(parentdict, childdict)
    ParentChildSku(wb1, parentdict, childdict)
        

def RRP1Lookup(wb1, search):
    ws2 = wb1.worksheets[7]
    mr2 = ws2.max_row
    
    # l = list(cell.value if cell.value is not None else "" for cell in ws2['O'])[2:]
    # l = [i for i in l if i]
    # # print(l)
    # if search in l:
    #     return ws2.cell(row = l.index(search)+3, column = 16).value
    # else:
    #     return "-"
    for i in range(3, mr2 + 1):
        if ws2.cell(row = i, column = 15).value==search:
            # print('here')
            return ws2.cell(row = i, column = 16).value
    return "-"

def WebLookup(wb1, search, sheet_number):
    ws2 = wb1.worksheets[sheet_number]
    mr2 = ws2.max_row
    if ws2.cell(row = 2, column = 1).value=="No data was returned" or mr2==2:
        return "-"
    for i in range(2, mr2 + 1):
        if ws2.cell(row = i, column = 5).value==search:
            return ws2.cell(row = i, column = 10).value
    return "-"

def RRP1PlusWeb(val1, val2):
    if val1=="-":
        if val2!="-":
            return val2
    else:
        if val2!="-":
            return val1+val2
        else:
            return val1
    return "-"

# def ParentChildSkuOld(wb1):
#     ws1 = wb1.worksheets[8]
#     ws2 = wb1.worksheets[0]

#     for i in range(2, ws1.max_row+1):
#     # for i in range(7, 9):
#         childind = 0
#         parentind = 0
#         for j in range(2, ws2.max_row+1):
#             if childind==0 and ws2.cell(row = j, column = 1).value == ws1.cell(row = i, column = 2).value:
#                 childind = j
#             if parentind ==0 and ws2.cell(row = j, column = 1).value == ws1.cell(row = i, column = 1).value:
#                 parentind = j
#             if childind!=0 and parentind!=0:
#                 break
#         if parentind!=0 and childind!=0:
#             while ws2.cell(row = parentind, column = 1).value == ws1.cell(row = i, column = 1).value:
#                 if ws2.cell(row = parentind, column = 3).value == ws2.cell(row = childind, column = 3).value:
#                     ws2.cell(row = parentind, column = 4).value += ws2.cell(row = childind, column = 4).value
#                     ws2.cell(row = parentind, column = 5).value += ws2.cell(row = childind, column = 5).value
#                     ws2.cell(row = parentind, column = 6).value += ws2.cell(row = childind, column = 6).value
#                     ws2.cell(row = parentind, column = 7).value += ws2.cell(row = childind, column = 7).value
#                     ws2.cell(row = parentind, column = 8).value += ws2.cell(row = childind, column = 8).value
#                     ws2.cell(row = parentind, column = 9).value += ws2.cell(row = childind, column = 9).value
#                     ws2.cell(row = parentind, column = 10).value += ws2.cell(row = childind, column = 10).value
#                     ws2.cell(row = parentind, column = 11).value += ws2.cell(row = childind, column = 11).value
#                     parentind += 1
#                     childind += 1
#                 elif ws2.cell(row = parentind, column = 3).value < ws2.cell(row = childind, column = 3).value:
#                     parentind += 1
#                 elif ws2.cell(row = parentind, column = 3).value > ws2.cell(row = childind, column = 3).value:
#                     childind += 1

def ParentChildSku(wb1, parentdict, childdict):
    ws1 = wb1.worksheets[8]
    ws2 = wb1.worksheets[0]
    ws3 = wb1.worksheets[10]

    for i in range(2, ws1.max_row+1):
    # for i in range(7, 9):
        parentind = parentdict[ws1.cell(row = i, column = 1).value][0]
        childind = childdict[ws1.cell(row = i, column = 2).value][0]
        oldparentind = parentdict[ws1.cell(row = i, column = 1).value][1]
        oldchildind = childdict[ws1.cell(row = i, column = 2).value][1]
        if parentind!=0 and childind!=0:
            while ws2.cell(row = parentind, column = 1).value == ws1.cell(row = i, column = 1).value:
                if ws2.cell(row = parentind, column = 3).value == ws2.cell(row = childind, column = 3).value:
                    # working
                    ws2.cell(row = parentind, column = 4).value += ws2.cell(row = childind, column = 4).value
                    ws2.cell(row = parentind, column = 5).value += ws2.cell(row = childind, column = 5).value
                    ws2.cell(row = parentind, column = 6).value += ws2.cell(row = childind, column = 6).value
                    ws2.cell(row = parentind, column = 7).value += ws2.cell(row = childind, column = 7).value
                    ws2.cell(row = parentind, column = 8).value += ws2.cell(row = childind, column = 8).value
                    ws2.cell(row = parentind, column = 9).value += ws2.cell(row = childind, column = 9).value
                    ws2.cell(row = parentind, column = 10).value += ws2.cell(row = childind, column = 10).value
                    ws2.cell(row = parentind, column = 11).value = ws2.cell(row = parentind, column = 4).value + ws2.cell(row = parentind, column = 6).value + ws2.cell(row = parentind, column = 8).value - ws2.cell(row = parentind, column = 10).value

                    # styling
                    ws2.cell(row = parentind, column = 4).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
                    ws2.cell(row = parentind, column = 6).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
                    ws2.cell(row = parentind, column = 8).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D8E4BC'))
                    ws2.cell(row = parentind, column = 10).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))

                    # old working
                if ws3.cell(row = oldparentind, column = 3).value == ws3.cell(row = oldchildind, column = 3).value:
                    ws3.cell(row = oldparentind, column = 6).value += ws3.cell(row = oldchildind, column = 6).value
                    ws3.cell(row = oldparentind, column = 7).value += ws3.cell(row = oldchildind, column = 7).value
                    ws3.cell(row = oldparentind, column = 8).value += ws3.cell(row = oldchildind, column = 8).value
                    ws3.cell(row = oldparentind, column = 9).value += ws3.cell(row = oldchildind, column = 9).value
                    ws3.cell(row = oldparentind, column = 10).value += ws3.cell(row = oldchildind, column = 10).value
                    ws3.cell(row = oldparentind, column = 11).value += ws3.cell(row = oldchildind, column = 11).value
                    ws3.cell(row = oldparentind, column = 12).value += ws3.cell(row = oldchildind, column = 12).value
                    ws3.cell(row = oldparentind, column = 14).value = ws3.cell(row = oldparentind, column = 6).value + ws3.cell(row = oldparentind, column = 8).value + ws3.cell(row = oldparentind, column = 10).value - ws3.cell(row = oldparentind, column = 12).value
                    parentind += 1
                    childind += 1
                    oldparentind += 1
                    oldchildind += 1
                elif ws2.cell(row = parentind, column = 3).value < ws2.cell(row = childind, column = 3).value:
                    parentind += 1
                    oldparentind += 1
                elif ws2.cell(row = parentind, column = 3).value > ws2.cell(row = childind, column = 3).value:
                    childind += 1
                    oldchildind += 1

def ParentChildWh(wb1, index, oldindex):
    ws1 = wb1.worksheets[12]
    ws2 = wb1.worksheets[0]
    ws3 = wb1.worksheets[10]
    for i in range(2, ws1.max_row + 1):
        if ws1.cell(row=i, column=1).value!=None:
            parentind = index
            oldparentind = oldindex
            while ws2.cell(row = parentind, column = 3).value != ws1.cell(row = i, column = 1).value:
                parentind -= 1
                oldparentind -= 1
            childind = parentind
            oldchildind = oldparentind
            if ws1.cell(row = i, column = 1).value > ws1.cell(row = i, column = 2).value:
                while ws2.cell(row = childind, column = 3).value != ws1.cell(row = i, column = 2).value:
                    # print(ws2.cell(row = childind, column = 3).value, ws1.cell(row = i, column = 2).value)
                    childind -= 1
                    oldchildind -= 1
            else:
                while ws2.cell(row = childind, column = 3).value < ws1.cell(row = i, column = 2).value and childind<index and oldchildind<oldindex:
                    # print(ws2.cell(row = childind, column = 3).value, ws1.cell(row = i, column = 2).value)
                    # print(val)
                    childind += 1
                    oldchildind += 1
            # working
            if ws2.cell(row = childind, column = 3).value == ws1.cell(row = i, column = 2).value and ws2.cell(row = parentind, column = 3).value == ws1.cell(row = i, column = 1).value:
                ws2.cell(row = parentind, column = 4).value += ws2.cell(row = childind, column = 4).value
                ws2.cell(row = parentind, column = 5).value += ws2.cell(row = childind, column = 5).value
                ws2.cell(row = parentind, column = 6).value += ws2.cell(row = childind, column = 6).value
                ws2.cell(row = parentind, column = 7).value += ws2.cell(row = childind, column = 7).value
                ws2.cell(row = parentind, column = 8).value += ws2.cell(row = childind, column = 8).value
                ws2.cell(row = parentind, column = 9).value += ws2.cell(row = childind, column = 9).value
                ws2.cell(row = parentind, column = 10).value += ws2.cell(row = childind, column = 10).value
                ws2.cell(row = parentind, column = 11).value = ws2.cell(row = parentind, column = 4).value + ws2.cell(row = parentind, column = 6).value + ws2.cell(row = parentind, column = 8).value - ws2.cell(row = parentind, column = 10).value

                # styling
                ws2.cell(row = parentind, column = 4).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00C5D9F1'))
                ws2.cell(row = parentind, column = 6).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='008DB4E2'))
                ws2.cell(row = parentind, column = 8).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00D8E4BC'))
                ws2.cell(row = parentind, column = 10).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor = openpyxl.styles.colors.Color(rgb='00BFBFBF'))

            # old working
            if ws3.cell(row = oldchildind, column = 3).value == ws1.cell(row = i, column = 2).value and ws3.cell(row = oldparentind, column = 3).value == ws1.cell(row = i, column = 1).value:
                ws3.cell(row = oldparentind, column = 6).value += ws3.cell(row = oldchildind, column = 6).value
                ws3.cell(row = oldparentind, column = 7).value += ws3.cell(row = oldchildind, column = 7).value
                ws3.cell(row = oldparentind, column = 8).value += ws3.cell(row = oldchildind, column = 8).value
                ws3.cell(row = oldparentind, column = 9).value += ws3.cell(row = oldchildind, column = 9).value
                ws3.cell(row = oldparentind, column = 10).value += ws3.cell(row = oldchildind, column = 10).value
                ws3.cell(row = oldparentind, column = 11).value += ws3.cell(row = oldchildind, column = 11).value
                ws3.cell(row = oldparentind, column = 12).value += ws3.cell(row = oldchildind, column = 12).value
                ws3.cell(row = oldparentind, column = 14).value = ws3.cell(row = oldparentind, column = 6).value + ws3.cell(row = oldparentind, column = 8).value + ws3.cell(row = oldparentind, column = 10).value - ws3.cell(row = oldparentind, column = 12).value

def startpoint(variant):
    print('Started')
    if variant=="paste":
        wb1 = load_workbook('report/input/Base File (Paste).xlsx')
    elif variant=="brush":
        wb1 = load_workbook('report/input/Base File (Brush).xlsx')
    else:
        wb1 = load_workbook('report/input/Base File (PCP).xlsx')
    ws1 = wb1.active

    # ws2 = wb1.worksheets[1]

    # ws3 = wb1.worksheets[2]
    SAPDump(wb1, variant)

    # WebSheets(wb1, "", "report/input/web_south.xlsx", 3)
    # WebSheets(wb1, "", "report/input/web_north.xlsx", 4)
    # WebSheets(wb1, "", "report/input/web_west.xlsx", 5)
    # WebSheets(wb1, "", "report/input/web_east.xlsx", 6)

    # web download
    WebSheets(wb1, "http://cpindia.win.colpal.com/hubli/scripts/dwnld_xls.asp", "report/input/web_south", 3)
    WebSheets(wb1, "http://cpindia.win.colpal.com/north/scripts/dwnld_xls.asp", "report/input/web_north", 4)
    WebSheets(wb1, "http://cpindia.win.colpal.com/despatch/scripts/dwnld_xls.asp", "report/input/web_west", 5)
    WebSheets(wb1, "http://cpindia.win.colpal.com/koldc/scripts/dwnld_xls.asp", "report/input/web_east", 6)

    RRP1Copy(wb1)

    SAPRRP1ToWorking(wb1)

    # ParentChildSku(wb1)
    for row in ws1.iter_rows(min_row= 1, max_row = ws1.max_row, min_col = 1, max_col = ws1.max_column):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    ws7 = wb1.worksheets[7]
    ws7.delete_cols(15)
    ws7.delete_cols(15)
    filename = datetime.date.today().strftime("%d%b%Y")
    if variant=="paste":
        wb1.save('report/output/'+filename+' (Paste).xlsx')
    elif variant=="brush":
        wb1.save('report/output/'+filename+' (Brush).xlsx')
    else:
        wb1.save('report/output/'+filename+' (PCP).xlsx')
    return True